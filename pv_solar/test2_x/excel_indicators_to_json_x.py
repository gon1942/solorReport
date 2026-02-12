#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import os
from datetime import datetime, date
from typing import Any, Dict, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


def is_visible_sheet(ws: Worksheet) -> bool:
    # openpyxl: 'visible', 'hidden', 'veryHidden'
    return getattr(ws, "sheet_state", "visible") == "visible"


def pick_active_visible_sheet(wb: Workbook) -> Worksheet:
    ws = wb.active
    if is_visible_sheet(ws):
        return ws

    for s in wb.worksheets:
        if is_visible_sheet(s):
            return s

    raise RuntimeError("보이는(visible) 시트가 없습니다.")


def is_visible_row(ws: Worksheet, row_idx: int) -> bool:
    rd = ws.row_dimensions.get(row_idx)
    if rd is None:
        return True
    return not bool(getattr(rd, "hidden", False))


def as_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x)


def normalize_excel_value(value: Any, field_type: Optional[Any] = None) -> Tuple[Any, Optional[str]]:
    """
    value: 셀 값 (openpyxl이 반환)
    field_type: 엑셀의 type 컬럼 값(문자열일 가능성이 큼)
    반환: (정규화된 값, 에러 메시지 or None)
    """
    ft = as_str(field_type).strip().lower()

    # None 그대로
    if value is None:
        return None, None

    # 이미 날짜/시간 타입이면 ISO로
    if isinstance(value, (datetime, date)):
        try:
            if isinstance(value, datetime):
                return value.date().isoformat(), None
            return value.isoformat(), None
        except Exception as e:
            return as_str(value), f"date_normalize_error: {e}"

    # 타입 힌트 기반 처리
    # 예: 'date (day/month/year)', 'years', 'months', 'MWp', 'Mwac' 등
    try:
        if "date" in ft:
            # 문자열 날짜일 수도 있어 간단 파싱 시도(실패 시 원문 유지)
            if isinstance(value, str):
                s = value.strip()
                # 흔한 포맷들만 가볍게 시도
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
                    try:
                        return datetime.strptime(s, fmt).date().isoformat(), None
                    except Exception:
                        pass
                return s, None
            return value, None

        if any(k in ft for k in ("number", "years", "months", "mwp", "mwac", "ratio", "capacity", "percent", "%")):
            # 숫자로 강제 변환 가능한 경우 float/int로
            if isinstance(value, (int, float)):
                return value, None
            if isinstance(value, str):
                s = value.strip().replace(",", "")
                if s == "":
                    return None, None
                # 퍼센트 문자열 "10%" 처리
                if s.endswith("%"):
                    s2 = s[:-1].strip()
                    try:
                        return float(s2) / 100.0, None
                    except Exception:
                        return s, None
                try:
                    # 정수/실수 구분
                    f = float(s)
                    if f.is_integer():
                        return int(f), None
                    return f, None
                except Exception:
                    return value, None
            return value, None

    except Exception as e:
        return value, f"normalize_error: {e}"

    # 기본: 문자열이면 trim, 그 외 그대로
    if isinstance(value, str):
        v = value.strip()
        return (v if v != "" else None), None

    return value, None


def is_section_header(label: Any, item_type: Any, value: Any) -> bool:
    if label is None:
        return False
    lbl = as_str(label).strip()
    if lbl == "":
        return False

    t = as_str(item_type).strip()
    v = as_str(value).strip() if value is not None else ""

    # type이 비었고 value도 비었으면 섹션으로 간주
    if t == "" and v == "":
        return True
    return False


def is_all_empty(label: Any, item_type: Any, value: Any) -> bool:
    return (as_str(label).strip() == "" and as_str(item_type).strip() == "" and (value is None or as_str(value).strip() == ""))


def parse_sheet(
    ws: Worksheet,
    label_col: int,
    type_col: int,
    value_col: int,
    start_row: int,
    default_section: str = "General inputs",
) -> Dict[str, list]:
    sections: Dict[str, list] = {}
    current_section = default_section
    sections.setdefault(current_section, [])

    # ws.max_row는 큰 값이 될 수 있으니, "완전 공백 행 연속"으로 조기 종료 옵션을 둘 수도 있으나
    # 여기서는 단순/명확하게 끝까지 스캔합니다.
    for row_idx in range(start_row, ws.max_row + 1):
        # ✅ 숨김 행 제외
        if not is_visible_row(ws, row_idx):
            continue

        label = ws.cell(row=row_idx, column=label_col).value
        item_type = ws.cell(row=row_idx, column=type_col).value
        value = ws.cell(row=row_idx, column=value_col).value

        # 섹션 헤더 판단 전 value 정규화(섹션은 value/type 비어있어야 하므로 안전)
        value_norm, err = normalize_excel_value(value, field_type=item_type)

        # 완전 공백 행은 스킵
        if is_all_empty(label, item_type, value_norm):
            continue

        if is_section_header(label, item_type, value_norm):
            current_section = as_str(label).strip()
            sections.setdefault(current_section, [])
            continue

        # 일반 항목
        sections.setdefault(current_section, [])
        item = {
            "label": as_str(label).strip() if label is not None else None,
            "row": row_idx,
            "type": as_str(item_type).strip() if item_type is not None else None,
            "value": value_norm,
        }
        if err:
            item["error"] = err

        sections[current_section].append(item)

    return sections


def parse_args():
    p = argparse.ArgumentParser(
        description="Excel 입력시트(지표/항목) → JSON 변환 (숨김 시트/숨김 행 제외)"
    )
    p.add_argument("xlsx", help="입력 엑셀 파일 경로")
    p.add_argument("-o", "--out", required=True, help="출력 JSON 경로")

    # 시트 지정 옵션(필요 시)
    p.add_argument("--sheet", default=None, help="(선택) 특정 시트명. 숨김 시트면 오류")

    # 컬럼 지정 (기본값은 흔한 Inputsheets 구조 가정: A=label, B=type, E=value 등)
    p.add_argument("--label-col", default="A", help="label 컬럼(예: A)")
    p.add_argument("--type-col", default="B", help="type 컬럼(예: B)")
    p.add_argument("--value-col", default="E", help="value 컬럼(예: E)")
    p.add_argument("--start-row", type=int, default=1, help="시작 행(기본 1)")
    p.add_argument("--default-section", default="General inputs", help="섹션 미지정 시 기본 섹션명")

    return p.parse_args()


def col_to_index(col: str) -> int:
    # 'A' -> 1, 'E' -> 5
    c = col.strip().upper()
    if not c.isalpha():
        raise ValueError(f"컬럼 문자는 A~Z 형식이어야 합니다: {col}")
    idx = 0
    for ch in c:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def main():
    args = parse_args()

    xlsx_path = os.path.abspath(args.xlsx)
    out_path = os.path.abspath(args.out)

    label_col = col_to_index(args.label_col)
    type_col = col_to_index(args.type_col)
    value_col = col_to_index(args.value_col)

    wb = load_workbook(xlsx_path, data_only=True)

    # ✅ "활성화된 시트만" 사용 (단, 숨김이면 첫 visible로 대체)
    if args.sheet:
        ws = wb[args.sheet]
        if not is_visible_sheet(ws):
            raise RuntimeError(f"요청한 시트가 숨김 상태입니다: {ws.title} (state={ws.sheet_state})")
    else:
        ws = pick_active_visible_sheet(wb)

    sections = parse_sheet(
        ws,
        label_col=label_col,
        type_col=type_col,
        value_col=value_col,
        start_row=args.start_row,
        default_section=args.default_section,
    )

    result = {
        "source": {
            "file": xlsx_path,
            "sheet": ws.title,
            "label_col": label_col,
            "label_col_letter": args.label_col.upper(),
            "type_col": type_col,
            "type_col_letter": args.type_col.upper(),
            "value_col": value_col,
            "value_col_letter": args.value_col.upper(),
            "start_row": args.start_row,
        },
        "sections": sections,
    }

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"[OK] Wrote: {out_path}")
    print(f"     Sheet: {ws.title} (state={ws.sheet_state})")


if __name__ == "__main__":
    main()
