#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel (Project 1~10) → JSON 변환 스크립트
- 시트에서 숨김(hidden) 처리된 '행'은 제외합니다.
- "Project 1" ~ "Project 10" 헤더를 자동 탐지하여 각 프로젝트별 값을 추출합니다.
- 기본 레이아웃(라벨=B열, 타입=C열, 노트=D열)을 가정하지만 옵션으로 변경 가능합니다.

사용 예:
  python excel_projects_to_json.py /path/to/file.xlsx -o projects.json
  python excel_projects_to_json.py /path/to/file.xlsx --sheet "INPUTSHEET(Batch2)" -o projects.json
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import os
import re
from collections import OrderedDict
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


PROJECT_RE = re.compile(r"^\s*Project\s*([1-9]|10)\s*$", re.IGNORECASE)


def is_hidden_row(ws, row_idx: int) -> bool:
    """엑셀에서 숨김 처리된 행인지 판단합니다."""
    dim = ws.row_dimensions.get(row_idx)
    if dim is None:
        return False
    if getattr(dim, "hidden", False):
        return True
    height = getattr(dim, "height", None)
    if height is not None and height == 0:
        return True
    return False


def normalize_cell_value(v: Any) -> Any:
    """JSON 직렬화가 잘 되도록 값 정규화."""
    if v is None:
        return None

    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    if isinstance(v, _dt.time):
        return v.strftime("%H:%M:%S")

    if isinstance(v, (int, float, bool)):
        return v

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # Excel error (#DIV/0!, #N/A 등)는 null 처리
        if s.startswith("#"):
            return None
        return s

    # 기타 타입은 문자열로
    return str(v)


def discover_project_columns(ws, max_row: int = 30, max_col: int = 80) -> List[Tuple[int, int, str, int]]:
    """
    시트에서 "Project 1~10" 헤더를 찾아 (project_no, col_idx, project_name, header_row) 리스트로 반환.
    각 프로젝트 번호별로 '가장 먼저 발견된' 열만 사용합니다(중복 헤더 방지).
    """
    found: Dict[int, Tuple[int, int, str]] = {}

    rmax = min(ws.max_row, max_row)
    cmax = min(ws.max_column, max_col)

    for r in range(1, rmax + 1):
        for c in range(1, cmax + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                m = PROJECT_RE.match(v)
                if m:
                    num = int(m.group(1))
                    # 동일 프로젝트 번호 중복 발견 시, 첫 발견(대개 가장 왼쪽)을 유지
                    found.setdefault(num, (r, c, v.strip()))

    cols: List[Tuple[int, int, str, int]] = []
    for num in range(1, 11):
        if num in found:
            r, c, name = found[num]
            cols.append((num, c, name, r))
    return cols


def parse_sheet(
    ws,
    label_col: int = 2,   # B
    type_col: int = 3,    # C
    note_col: int = 4,    # D
    detect_rows: int = 30,
    detect_cols: int = 80,
) -> Dict[str, OrderedDict]:
    """
    시트를 파싱하여 { "Project 1": OrderedDict({section: [items...] }), ... } 형태로 반환
    """
    project_cols = discover_project_columns(ws, max_row=detect_rows, max_col=detect_cols)
    if not project_cols:
        raise ValueError(f'시트 "{ws.title}"에서 "Project 1~10" 헤더를 찾지 못했습니다.')

    proj_names = [name for _, _, name, _ in project_cols]
    proj_col_indices = [col for _, col, _, _ in project_cols]

    projects: Dict[str, OrderedDict] = {name: OrderedDict() for name in proj_names}

    current_section = "Ungrouped"
    for p in projects.values():
        p[current_section] = []

    for r in range(1, ws.max_row + 1):
        if is_hidden_row(ws, r):
            continue

        label_raw = ws.cell(r, label_col).value
        if label_raw is None:
            continue

        label = label_raw.strip() if isinstance(label_raw, str) else str(label_raw).strip()
        if not label:
            continue

        # 섹션 헤더 판단:
        # - 라벨은 문자열
        # - 프로젝트 값 열들은 전부 None
        # - type/note 도 비어있음
        proj_vals = [ws.cell(r, c).value for c in proj_col_indices]
        type_raw = ws.cell(r, type_col).value
        note_raw = ws.cell(r, note_col).value

        def _is_empty_meta(x: Any) -> bool:
            return x is None or (isinstance(x, str) and not x.strip())

        if isinstance(label_raw, str) and all(v is None for v in proj_vals) and _is_empty_meta(type_raw) and _is_empty_meta(note_raw):
            current_section = label
            for p in projects.values():
                p.setdefault(current_section, [])
            continue

        t = normalize_cell_value(type_raw)
        n = normalize_cell_value(note_raw)

        for name, c in zip(proj_names, proj_col_indices):
            v = normalize_cell_value(ws.cell(r, c).value)
            item: Dict[str, Any] = {"label": label, "value": v, "row": r}
            if t is not None:
                item["type"] = t
            if n is not None:
                item["note"] = n
            projects[name][current_section].append(item)

    # 빈 섹션 제거
    for pname in list(projects.keys()):
        secmap = projects[pname]
        for sec in list(secmap.keys()):
            if not secmap[sec]:
                del secmap[sec]

    return projects


def pick_default_sheet(wb) -> str:
    """Project 헤더가 있는 첫 번째 시트를 기본으로 선택."""
    for s in wb.sheetnames:
        ws = wb[s]
        if discover_project_columns(ws):
            return s
    return wb.sheetnames[0]


def build_output(
    xlsx_path: str,
    sheet_name: str,
    projects: Dict[str, OrderedDict],
    label_col: int,
    type_col: int,
    note_col: int,
) -> Dict[str, Any]:
    project_cols = discover_project_columns(load_workbook(xlsx_path, data_only=True)[sheet_name])
    proj_map = {name: get_column_letter(col) for _, col, name, _ in project_cols}

    return {
        "source": {
            "file": os.path.abspath(xlsx_path),
            "sheet": sheet_name,
            "label_col": get_column_letter(label_col),
            "type_col": get_column_letter(type_col),
            "note_col": get_column_letter(note_col),
            "project_columns": proj_map,
            "excluded_hidden_rows": True,
        },
        "projects": projects,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="입력 엑셀 파일(.xlsx)")
    ap.add_argument("-o", "--output", required=True, help="출력 JSON 경로")
    ap.add_argument("--sheet", help="시트명(미지정 시 Project 헤더가 있는 첫 시트 자동 선택)")
    ap.add_argument("--label-col", default="B", help="라벨 컬럼(예: A, B, C). 기본=B")
    ap.add_argument("--type-col", default="C", help="타입/형식 컬럼(예: C). 기본=C")
    ap.add_argument("--note-col", default="D", help="노트/설명 컬럼(예: D). 기본=D")
    ap.add_argument("--detect-rows", type=int, default=30, help="Project 헤더 탐지 최대 행(기본 30)")
    ap.add_argument("--detect-cols", type=int, default=80, help="Project 헤더 탐지 최대 열(기본 80)")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    sheet_name = args.sheet or pick_default_sheet(wb)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f'시트 "{sheet_name}"를 찾을 수 없습니다. 가능한 시트: {wb.sheetnames}')

    ws = wb[sheet_name]

    def col_to_idx(col_letter: str) -> int:
        col_letter = col_letter.strip().upper()
        if not re.fullmatch(r"[A-Z]{1,3}", col_letter):
            raise ValueError(f"컬럼 지정이 올바르지 않습니다: {col_letter}")
        idx = 0
        for ch in col_letter:
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
        return idx

    label_col = col_to_idx(args.label_col)
    type_col = col_to_idx(args.type_col)
    note_col = col_to_idx(args.note_col)

    projects = parse_sheet(
        ws,
        label_col=label_col,
        type_col=type_col,
        note_col=note_col,
        detect_rows=args.detect_rows,
        detect_cols=args.detect_cols,
    )

    out = build_output(args.xlsx, sheet_name, projects, label_col, type_col, note_col)

    os.makedirs(os.path.dirname(os.path.abspath(args.output)) or ".", exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"[OK] JSON 생성 완료: {args.output}")
    print(f"     sheet={sheet_name}, projects={len(projects)}")


if __name__ == "__main__":
    main()

"""
python excel_indicators_to_json_v1.py "/work/11.airun/solar/workspaces_pv_solar/pv_solar_sample_docs/지상형태양광_Inputsheet.xlsx" \
  --sheet "INPUTSHEET(Batch2)" \
  -o "projects_1_10.json"

"""