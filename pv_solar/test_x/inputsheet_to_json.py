#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
InputSheet Excel -> JSON extractor (CAPEX/DEVEX Approval workflow)

목적
- Input sheet 엑셀에서 'Project 1' 같은 특정 프로젝트 컬럼의 값을 뽑아
  LLM 입력에 적합한 JSON으로 정규화합니다.

특징
- --project "Project 1" (헤더에서 프로젝트 컬럼 자동 탐색) 또는 --col E (컬럼 강제 지정)
- (일반적으로) 왼쪽에는 항목명(label), 오른쪽 프로젝트 컬럼에는 값(value)이 있는 형태를 가정
- 섹션(General inputs, CAPEX, Development cost 등) 단위로 라벨/값을 수집하여 JSON으로 출력

사용 예
  python inputsheet_to_json.py "/work/proj/airun/002_poc/airun/pv_solar_sample_docs/지상형태양광_Inputsheet.xlsx" --project "Project 1" -o inputs_project1.json

  python inputsheet_to_json.py "/work/proj/airun/002_poc/airun/pv_solar_sample_docs/지상형태양광_Inputsheet.xlsx" --col E -o inputs_project1.json
  


주의
- InputSheet 레이아웃이 다르면, find_project_column(), extract_kv_pairs()의 탐색 규칙을 조정해야 합니다.
"""


import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import date, datetime
from decimal import Decimal

def norm_text(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def is_blank(x: Any) -> bool:
    return norm_text(x) == ""


def cell_val(ws: Worksheet, r: int, c: int) -> Any:
    return ws.cell(row=r, column=c).value


def find_project_column(ws: Worksheet, project_name: str, scan_rows: int = 30, scan_cols: int = 30) -> Optional[int]:
    """
    시트 상단 영역에서 project_name 텍스트가 들어있는 셀을 찾고 그 열 번호를 반환합니다.
    """
    target = project_name.strip().lower()
    for r in range(1, scan_rows + 1):
        for c in range(1, scan_cols + 1):
            v = norm_text(cell_val(ws, r, c)).lower()
            if v == target:
                return c
    # 부분 매칭(예: "Project 1 (PV)" 등)
    for r in range(1, scan_rows + 1):
        for c in range(1, scan_cols + 1):
            v = norm_text(cell_val(ws, r, c)).lower()
            if target and target in v:
                return c
    return None


def col_letter_to_index(col_letter: str) -> int:
    col_letter = col_letter.strip().upper()
    if not re.fullmatch(r"[A-Z]+", col_letter):
        raise ValueError(f"Invalid column letter: {col_letter}")
    idx = 0
    for ch in col_letter:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def detect_label_cell(ws: Worksheet, r: int, label_cols: List[int]) -> Tuple[str, Optional[int]]:
    """
    해당 row에서 label이 들어있는 셀을 찾습니다.
    기본적으로 A~D 같은 왼쪽 영역을 label 후보로 보고,
    가장 먼저 발견되는 '의미 있는 문자열'을 label로 간주합니다.
    """
    for c in label_cols:
        v = norm_text(cell_val(ws, r, c))
        # 너무 짧은 값(예: "-", "1" 등) 제외
        if v and len(v) >= 2:
            return v, c
    return "", None


def looks_like_section_title(label: str) -> bool:
    """
    섹션 타이틀로 보이는지 간단 판정:
    - 'General inputs', 'CAPEX', 'Development cost' 등
    - 끝이 ':' 인 경우
    - 대문자 비율이 높거나(짧은 단어) 단독 키워드
    """
    s = label.strip()
    if not s:
        return False
    if s.endswith(":"):
        return True

    key = s.lower()
    section_keywords = [
        "general inputs", "general", "capex", "devex", "development cost",
        "opex", "revenues", "schedule", "timeline", "assumptions"
    ]
    if any(k == key for k in section_keywords):
        return True
    if any(k in key for k in ["general inputs", "development", "capex", "devex"]):
        return True

    # 단어 수가 적고(1~3), 알파벳 위주면 섹션일 가능성
    words = re.split(r"\s+", s)
    if 1 <= len(words) <= 3:
        alpha = sum(ch.isalpha() for ch in s)
        if alpha / max(1, len(s)) > 0.6 and s[0].isupper():
            return True

    return False


def extract_kv_pairs(
    ws: Worksheet,
    value_col: int,
    label_cols: List[int] = None,
    start_row: int = 1,
    end_row: Optional[int] = None
) -> List[Dict[str, Any]]:
    """
    시트에서 (section, label, value) 형태로 쭉 수집합니다.
    """
    if label_cols is None:
        label_cols = [1, 2, 3, 4]  # A~D

    max_row = ws.max_row
    if end_row is None or end_row > max_row:
        end_row = max_row

    rows: List[Dict[str, Any]] = []
    current_section = "Uncategorized"

    for r in range(start_row, end_row + 1):
        label, label_col_found = detect_label_cell(ws, r, label_cols)
        val = cell_val(ws, r, value_col)

        label_n = norm_text(label)
        val_n = norm_text(val)

        # 행이 완전히 비면 스킵
        if not label_n and not val_n:
            continue

        # 섹션 타이틀 갱신(값 칸이 비어있고 label이 섹션처럼 보이면)
        if label_n and looks_like_section_title(label_n) and is_blank(val):
            current_section = label_n.rstrip(":").strip()
            continue

        # label이 없고 값만 있는 경우는 거의 의미 없으니 스킵
        if not label_n:
            continue

        rows.append({
            "section": current_section,
            "label": label_n,
            "value": val if val_n != "" else None,
            "row": r,
            "label_col": label_col_found,
            "value_col": value_col,
        })

    return rows


def group_by_section(items: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    out: Dict[str, List[Dict[str, Any]]] = {}
    for it in items:
        sec = it.get("section") or "Uncategorized"
        out.setdefault(sec, []).append({
            "label": it.get("label"),
            "value": it.get("value"),
            "row": it.get("row")
        })
    return out


def pick_sheet(wb, preferred: Optional[str] = None) -> Worksheet:
    if preferred:
        if preferred in wb.sheetnames:
            return wb[preferred]
        # 부분 매칭
        low = preferred.lower()
        for n in wb.sheetnames:
            if low in n.lower():
                return wb[n]
    return wb[wb.sheetnames[0]]


def main():
    ap = argparse.ArgumentParser(description="Extract InputSheet values to JSON.")
    ap.add_argument("xlsx", help="Input Excel (.xlsx)")
    ap.add_argument("--sheet", default=None, help="Sheet name (optional). Default: first sheet.")
    ap.add_argument("--project", default=None, help='Project column header text, e.g. "Project 1"')
    ap.add_argument("--col", default=None, help="Value column letter, e.g. E (overrides --project)")
    ap.add_argument("-o", "--out", default=None, help="Output JSON path. Default: <xlsx_stem>.json")
    ap.add_argument("--start-row", type=int, default=1, help="Start row for scanning")
    ap.add_argument("--end-row", type=int, default=0, help="End row for scanning (0=auto)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        raise SystemExit(f"[ERROR] not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = pick_sheet(wb, args.sheet)

    # value column 결정
    value_col: Optional[int] = None
    if args.col:
        value_col = col_letter_to_index(args.col)
    elif args.project:
        value_col = find_project_column(ws, args.project)
        if value_col is None:
            raise SystemExit(f"[ERROR] cannot find project column header: {args.project}")
    else:
        # 아무 것도 없으면 E를 기본 가정(많이 쓰는 패턴)
        value_col = col_letter_to_index("E")

    end_row = args.end_row if args.end_row and args.end_row > 0 else ws.max_row

    items = extract_kv_pairs(
        ws,
        value_col=value_col,
        label_cols=[1, 2, 3, 4],  # A~D
        start_row=args.start_row,
        end_row=end_row
    )

    by_section = group_by_section(items)

    out_obj = {
        "source": {
            "file": str(xlsx_path),
            "sheet": ws.title,
            "project": args.project,
            "value_col": value_col,
            "value_col_letter": None,
        },
        "sections": by_section,
        "flat": items,
    }

    # value_col -> letter
    def idx_to_letter(idx: int) -> str:
        s = ""
        while idx > 0:
            idx, rem = divmod(idx - 1, 26)
            s = chr(rem + ord("A")) + s
        return s

    out_obj["source"]["value_col_letter"] = idx_to_letter(value_col)

    out_path = Path(args.out) if args.out else xlsx_path.with_suffix("").with_name(xlsx_path.stem + ".json")
    out_path = out_path.expanduser().resolve()
    # out_path.write_text(json.dumps(out_obj, ensure_ascii=False, indent=2), encoding="utf-8")
    def _json_default(o):
        if isinstance(o, (datetime, date)):
            return o.isoformat()
        if isinstance(o, Decimal):
            return float(o)
        # openpyxl가 간혹 numpy 타입/기타를 줄 수 있어 문자열 fallback
        return str(o)

    out_path.write_text(
        json.dumps(out_obj, ensure_ascii=False, indent=2, default=_json_default),
        encoding="utf-8"
    )

    

    print(f"[OK] sheet={ws.title} value_col={out_obj['source']['value_col_letter']} -> {out_path}")
    print(f"[OK] sections={len(by_section)} rows={len(items)}")


if __name__ == "__main__":
    main()
