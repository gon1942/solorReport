#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import re
import datetime
from pathlib import Path
from typing import Any, Dict, Tuple, Optional, List

import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string

PROJECT_RE = re.compile(r"^Project\s*\d+\s*$", re.IGNORECASE)

# Top-level 섹션명 후보(섹션 끝 경계 판단용). 필요 시 추가하세요.
TOP_LEVEL_SECTIONS = [
    "General inputs",
    "Production",
    "Opex - year 1",
    "Opex - year 2",
    "CAPEX",
    "OPEX",
    "Revenues",
    "Financing",
    "Debt",
    "Equity",
    "Assumptions",
]

# Opex 표에서 자주 등장하는 필드 라벨(카테고리 헤더 판별용)
OPEX_FIELD_LABELS = {
    "currency",
    "cost",
    "indexed?",
    "indexation reference date",
    "indexation %",
    "cost starting date",
    "comments",
    "hectares",
    "unit",
}

# Excel 0-date 계열(의미 없는 날짜) 방지
EXCEL_ZERO_DATE = datetime.date(1899, 12, 30)

# Production에서 그룹 하위(예: P50/P75/P90)로 넣고 싶은 라벨 패턴
P_LABEL_RE = re.compile(r"^P\d+\s*$", re.IGNORECASE)


def norm_text(v: Any) -> Optional[str]:
    if not isinstance(v, str):
        return None
    s = v.strip()
    return s if s else None


def to_json_value(v: Any) -> Any:
    """
    openpyxl 값 -> JSON 직렬화 가능한 값으로 변환
    - datetime/date/time 모두 ISO 문자열로 변환
    - Excel 0-date는 ""로 처리
    """
    if v is None:
        return None

    if isinstance(v, datetime.datetime):
        return v.isoformat()

    if isinstance(v, datetime.date):
        if v == EXCEL_ZERO_DATE:
            return ""
        return v.isoformat()

    if isinstance(v, datetime.time):
        return v.isoformat()

    if isinstance(v, float) and v.is_integer():
        return int(v)

    if isinstance(v, str):
        return v.strip()

    return v


def blank_if_empty(v: Any) -> Any:
    """
    빈값 처리 규칙:
    - None -> ""
    - "-" -> ""
    - "" 유지
    """
    if v is None:
        return ""
    if isinstance(v, str):
        s = v.strip()
        if s == "-" or s == "":
            return ""
        return s
    return v


def normalize_indexed_value(v: Any) -> str:
    """
    Indexed? 값 정규화:
    - 1 / "1" / True  -> "Yes"
    - 0 / "0" / False -> "No"
    - "Yes"/"No" 형태면 그대로
    - 빈값이면 ""
    """
    v = blank_if_empty(v)
    if v == "":
        return ""

    if isinstance(v, str):
        s = v.strip().lower()
        if s in ("yes", "y", "true"):
            return "Yes"
        if s in ("no", "n", "false"):
            return "No"
        if s == "1":
            return "Yes"
        if s == "0":
            return "No"
        return v.strip()

    if isinstance(v, bool):
        return "Yes" if v else "No"

    if isinstance(v, (int, float)):
        try:
            return "Yes" if int(v) == 1 else "No"
        except Exception:
            return str(v)

    return str(v)


def is_column_visible(ws, col_idx: int) -> bool:
    """
    '눈에 보이는 열' 판별:
    - column_dimensions.hidden == True 또는 width == 0이면 제외
    """
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    if dim is None:
        return True

    if getattr(dim, "hidden", False):
        return False

    width = getattr(dim, "width", None)
    if width is not None:
        try:
            if float(width) == 0.0:
                return False
        except Exception:
            pass

    return True


def find_project_header_row_visible(
    ws,
    scan_max_rows: int = 300,
) -> Tuple[int, Dict[int, str], List[Dict[str, str]], List[Dict[str, str]]]:
    """
    - Project 헤더 행을 찾고
    - '보이는 열'의 Project 컬럼만 반환
    - 같은 Project명이 중복이면 뒤쪽 제거

    반환:
      (header_row,
       project_cols_map: {col_idx: "Project n"},
       hidden_project_cols: [{"project":..., "col":...}],
       dropped_duplicates: [{"project":..., "col":..., "reason":...}])
    """
    best_row: Optional[int] = None
    best_visible: Dict[int, str] = {}
    best_hidden: Dict[int, str] = {}

    max_rows = min(ws.max_row, scan_max_rows)
    for r in range(1, max_rows + 1):
        row_visible: Dict[int, str] = {}
        row_hidden: Dict[int, str] = {}

        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and PROJECT_RE.match(v.strip()):
                if is_column_visible(ws, c):
                    row_visible[c] = v.strip()
                else:
                    row_hidden[c] = v.strip()

        if row_visible:
            if best_row is None or len(row_visible) > len(best_visible):
                best_row = r
                best_visible = row_visible
                best_hidden = row_hidden

    if best_row is None:
        raise RuntimeError("Project 헤더 행(Project 1, Project 2, ...)을 찾지 못했습니다.")

    # 연속 블록만 유지(오탐 방지)
    cols = sorted(best_visible.keys())
    contiguous = [cols[0]]
    for i in range(1, len(cols)):
        if cols[i] == cols[i - 1] + 1:
            contiguous.append(cols[i])
        else:
            break

    # 중복 Project명 제거(뒤쪽 제거)
    seen = set()
    dedup_map: Dict[int, str] = {}
    dropped_duplicates: List[Dict[str, str]] = []
    for c in contiguous:
        name = best_visible[c]
        if name in seen:
            dropped_duplicates.append(
                {"project": name, "col": get_column_letter(c), "reason": "duplicate_project_name"}
            )
            continue
        seen.add(name)
        dedup_map[c] = name

    hidden_project_cols = [
        {"project": best_hidden[c], "col": get_column_letter(c)}
        for c in sorted(best_hidden.keys())
    ]

    return best_row, dedup_map, hidden_project_cols, dropped_duplicates


def find_section_block_by_top_sections(
    ws,
    label_col: int,
    section_name: str,
    top_sections: List[str],
) -> Optional[Tuple[int, int]]:
    """
    섹션 범위를 '다음 Top-level 섹션명'이 등장할 때까지로 계산합니다.
    (Production 내부 소제목은 섹션 종료로 보면 안 되므로 이 방식을 사용)
    """
    target = section_name.strip().lower()
    top_set = {s.strip().lower() for s in top_sections if s and s.strip()}

    start_header: Optional[int] = None
    for r in range(1, ws.max_row + 1):
        v = norm_text(ws.cell(r, label_col).value)
        if v and v.lower() == target:
            start_header = r
            break

    if start_header is None:
        return None

    end = ws.max_row + 1
    for r in range(start_header + 1, ws.max_row + 1):
        v = norm_text(ws.cell(r, label_col).value)
        if v and v.lower() in top_set and v.lower() != target:
            end = r
            break

    return start_header + 1, end


def is_group_header_row(ws, r: int, label_col: int, type_col: int, value_col: int) -> bool:
    """
    Production 내부 소제목(그룹 헤더) 판별:
    - label 있음
    - type_col 비어있음
    - value_col 비어있음
    """
    label = norm_text(ws.cell(r, label_col).value)
    if not label:
        return False

    t = ws.cell(r, type_col).value
    if t is not None and norm_text(t):
        return False

    if ws.cell(r, value_col).value is not None:
        return False

    return True


def extract_section_kv_for_project_flat(
    ws,
    start_row: int,
    end_row: int,
    label_col: int,
    type_col: int,
    value_col: int,
    include_meta: bool,
    include_empty_as_blank: bool,
) -> Dict[str, Any]:
    """
    General inputs처럼 단순 Key-Value 추출
    - 빈값은 ""로 저장 가능
    """
    out: Dict[str, Any] = {}
    for r in range(start_row, end_row):
        key = norm_text(ws.cell(r, label_col).value)
        if not key:
            continue

        val = to_json_value(ws.cell(r, value_col).value)
        if include_empty_as_blank:
            val = blank_if_empty(val)

        if include_meta:
            tval = to_json_value(ws.cell(r, type_col).value)
            if include_empty_as_blank:
                tval = blank_if_empty(tval)
            out[key] = {
                "type": tval,
                "value": val,
                "row": r,
                "col": value_col,
                "col_letter": get_column_letter(value_col),
            }
        else:
            out[key] = val

    return out


def extract_section_kv_for_project_grouped(
    ws,
    start_row: int,
    end_row: int,
    label_col: int,
    type_col: int,
    value_col: int,
    include_meta: bool,
    include_empty_as_blank: bool,
) -> Dict[str, Any]:
    """
    Production처럼 소제목(그룹)이 있는 섹션을 그룹화하여 추출
    - "P50/P75/P90" 같은 행은 직전 그룹에 넣고,
      그 외 항목(Availability 등)은 그룹을 끊고 최상위로 저장
    - 빈값은 ""로 저장
    """
    out: Dict[str, Any] = {}
    current_group: Optional[str] = None

    for r in range(start_row, end_row):
        label = norm_text(ws.cell(r, label_col).value)
        if not label:
            continue

        # 그룹 헤더
        if is_group_header_row(ws, r, label_col, type_col, value_col):
            current_group = label
            out.setdefault(current_group, {})
            continue

        # 그룹 유지 규칙: Pxx 라벨일 때만 current_group 유지
        if current_group is not None and not P_LABEL_RE.match(label):
            current_group = None

        val = to_json_value(ws.cell(r, value_col).value)
        if include_empty_as_blank:
            val = blank_if_empty(val)

        if current_group:
            grp = out.setdefault(current_group, {})
            if include_meta:
                tval = to_json_value(ws.cell(r, type_col).value)
                if include_empty_as_blank:
                    tval = blank_if_empty(tval)
                grp[label] = {
                    "type": tval,
                    "value": val,
                    "row": r,
                    "col": value_col,
                    "col_letter": get_column_letter(value_col),
                }
            else:
                grp[label] = val
        else:
            if include_meta:
                tval = to_json_value(ws.cell(r, type_col).value)
                if include_empty_as_blank:
                    tval = blank_if_empty(tval)
                out[label] = {
                    "type": tval,
                    "value": val,
                    "row": r,
                    "col": value_col,
                    "col_letter": get_column_letter(value_col),
                }
            else:
                out[label] = val

    return out


def extract_gi_and_production_from_active_sheet(
    xlsx_path: str,
    label_col_letter: str = "B",
    include_meta: bool = False,
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active  # 활성 시트만

    header_row, project_cols_map, hidden_project_cols, dropped_duplicates = find_project_header_row_visible(ws)
    project_col_indices = sorted(project_cols_map.keys())
    if not project_col_indices:
        raise RuntimeError("보이는(Project 헤더가 있는) 프로젝트 열을 찾지 못했습니다.")

    label_col = column_index_from_string(label_col_letter)
    type_col = label_col + 1

    gi_block = find_section_block_by_top_sections(ws, label_col, "General inputs", TOP_LEVEL_SECTIONS)
    prod_block = find_section_block_by_top_sections(ws, label_col, "Production", TOP_LEVEL_SECTIONS)

    missing_sections = []
    if gi_block is None:
        missing_sections.append("General inputs")
    if prod_block is None:
        missing_sections.append("Production")

    projects: List[Dict[str, Any]] = []
    for col_idx in project_col_indices:
        pname = project_cols_map[col_idx]
        col_letter = get_column_letter(col_idx)

        general_inputs: Dict[str, Any] = {}
        production: Dict[str, Any] = {}

        if gi_block is not None:
            gi_start, gi_end = gi_block
            general_inputs = extract_section_kv_for_project_flat(
                ws,
                gi_start,
                gi_end,
                label_col,
                type_col,
                col_idx,
                include_meta=include_meta,
                include_empty_as_blank=True,  # 빈값은 "" 처리
            )

        if prod_block is not None:
            pr_start, pr_end = prod_block
            production = extract_section_kv_for_project_grouped(
                ws,
                pr_start,
                pr_end,
                label_col,
                type_col,
                col_idx,
                include_meta=include_meta,
                include_empty_as_blank=True,  # 모든 항목 + 빈값 "" 처리
            )

        projects.append(
            {
                "name": pname,
                "col": col_letter,
                "general_inputs": general_inputs,
                "production": production,
            }
        )

    return {
        "source": {
            "file": xlsx_path,
            "sheet": ws.title,
            "project_header_row": header_row,
            "label_col": label_col_letter,
            "projects_detected": [project_cols_map[c] for c in project_col_indices],
            "project_columns": [{"project": project_cols_map[c], "col": get_column_letter(c)} for c in project_col_indices],
            "hidden_project_columns": hidden_project_cols,
            "dropped_duplicate_projects": dropped_duplicates,
            "missing_sections": missing_sections,
        },
        "projects": projects,
    }


# -----------------------------
# Opex - year 1 추출
# -----------------------------

def find_opex_year1_sheet(wb) -> Optional[Any]:
    """
    'Opex - year 1' 시트를 자동 탐색
    1) 시트명에 'opex'와 'year'와 '1' 포함
    2) 시트 내 셀에서 'Opex - year 1' 텍스트 검색
    """
    for ws in wb.worksheets:
        name = (ws.title or "").lower()
        if "opex" in name and "year" in name and "1" in name:
            return ws

    target = "opex - year 1"
    for ws in wb.worksheets:
        max_rows = min(ws.max_row, 200)
        max_cols = min(ws.max_column, 60)
        for r in range(1, max_rows + 1):
            for c in range(1, max_cols + 1):
                v = norm_text(ws.cell(r, c).value)
                if v and v.lower() == target:
                    return ws

    return None


def find_opex_title_row_and_unit_col(ws) -> Tuple[int, int, int]:
    """
    Opex - year 1 표의 제목 행/라벨열/Unit 열을 찾습니다.
    반환: (title_row, label_col, unit_col)
    """
    target = "opex - year 1"
    max_rows = min(ws.max_row, 200)
    max_cols = min(ws.max_column, 80)

    title_row = 1
    label_col = 1

    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            v = norm_text(ws.cell(r, c).value)
            if v and v.lower() == target:
                title_row = r
                label_col = c
                break
        else:
            continue
        break

    unit_col = -1
    for c in range(label_col, max_cols + 1):
        v = norm_text(ws.cell(title_row, c).value)
        if v and v.lower() == "unit":
            unit_col = c
            break

    if unit_col == -1:
        unit_col = label_col + 2  # fallback

    return title_row, label_col, unit_col


def build_project_cols_for_opex_sheet(ws, title_row: int, unit_col: int) -> Tuple[Dict[int, str], List[Dict[str, str]]]:
    """
    Opex 시트에서 프로젝트 열 매핑 구성
    1) Project 1..n 헤더 행이 있으면 사용(보이는 열만)
    2) 없으면 title_row의 값 열들을 Project 1..n로 생성(보이는 열만)
    """
    try:
        _, project_cols_map, _, _ = find_project_header_row_visible(ws)
        cols = sorted(project_cols_map.keys())
        return project_cols_map, [{"project": project_cols_map[c], "col": get_column_letter(c)} for c in cols]
    except Exception:
        project_cols_map: Dict[int, str] = {}
        started = False
        idx = 1
        for c in range(unit_col + 1, ws.max_column + 1):
            if not is_column_visible(ws, c):
                continue
            txt = norm_text(ws.cell(title_row, c).value)
            if txt is None:
                if started:
                    break
                continue
            started = True
            project_cols_map[c] = f"Project {idx}"
            idx += 1

        cols = sorted(project_cols_map.keys())
        return project_cols_map, [{"project": project_cols_map[c], "col": get_column_letter(c)} for c in cols]


def is_opex_category_header(ws, r: int, label_col: int) -> bool:
    """
    Opex 카테고리 헤더 판별:
    - label 문자열 존재
    - label이 필드 라벨(OPEX_FIELD_LABELS)이 아님
    - 아래 1~5행 내에 Currency/Cost 등의 필드 라벨이 등장하면 카테고리로 판단
    """
    label = norm_text(ws.cell(r, label_col).value)
    if not label:
        return False

    if label.strip().lower() in OPEX_FIELD_LABELS:
        return False

    for rr in range(r + 1, min(r + 6, ws.max_row + 1)):
        nxt = norm_text(ws.cell(rr, label_col).value)
        if nxt and nxt.strip().lower() in OPEX_FIELD_LABELS:
            return True

    return False


def extract_opex_year1_from_sheet(
    xlsx_path: str,
    include_meta: bool = False,
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = find_opex_year1_sheet(wb)
    if ws is None:
        raise RuntimeError("Opex - year 1 시트를 찾지 못했습니다(시트명/내용에서 미탐지).")

    title_row, label_col, unit_col = find_opex_title_row_and_unit_col(ws)
    prompt_col = label_col + 1

    project_cols_map, project_columns_list = build_project_cols_for_opex_sheet(ws, title_row, unit_col)
    project_col_indices = sorted(project_cols_map.keys())
    if not project_col_indices:
        raise RuntimeError("Opex - year 1에서 프로젝트 값 열을 찾지 못했습니다.")

    # 프로젝트별 결과
    per_project: Dict[int, Dict[str, Any]] = {}
    for c in project_col_indices:
        per_project[c] = {
            "name": project_cols_map[c],
            "col": get_column_letter(c),
            "opex_year1": {},
        }

    current_category: Optional[str] = None

    for r in range(title_row + 1, ws.max_row + 1):
        label = norm_text(ws.cell(r, label_col).value)
        if not label:
            continue

        # 카테고리 헤더
        if is_opex_category_header(ws, r, label_col):
            current_category = label
            for c in project_col_indices:
                per_project[c]["opex_year1"].setdefault(current_category, {})

            # 카테고리 헤더 행에 값(예: 1)이 있으면 enabled로 저장 (빈값은 "")
            any_val = False
            header_vals: Dict[int, Any] = {}
            for c in project_col_indices:
                v = blank_if_empty(to_json_value(ws.cell(r, c).value))
                header_vals[c] = v
                if v != "":
                    any_val = True

            if any_val:
                for c in project_col_indices:
                    per_project[c]["opex_year1"][current_category]["enabled"] = header_vals[c]

            continue

        if current_category is None:
            current_category = "Opex - year 1"
            for c in project_col_indices:
                per_project[c]["opex_year1"].setdefault(current_category, {})

        key = label
        key_lower = key.strip().lower()

        unit_val = blank_if_empty(to_json_value(ws.cell(r, unit_col).value))
        prompt_val = blank_if_empty(to_json_value(ws.cell(r, prompt_col).value))

        # 프로젝트별 값을 먼저 수집
        row_vals: Dict[int, Any] = {}
        for c in project_col_indices:
            v = blank_if_empty(to_json_value(ws.cell(r, c).value))

            # Indexed? 0/1 -> No/Yes
            if key_lower == "indexed?":
                v = normalize_indexed_value(v)

            row_vals[c] = v

        # Comments: 프로젝트별 값이 하나라도 있으면 그걸 사용, 전부 비었으면 prompt 사용
        if key_lower == "comments":
            any_project_comment = any(row_vals[c] != "" for c in project_col_indices)
            final_comment_prompt = prompt_val if (not any_project_comment and prompt_val != "") else ""

            for c in project_col_indices:
                final_val = row_vals[c] if any_project_comment else (final_comment_prompt or "")
                if include_meta:
                    per_project[c]["opex_year1"][current_category][key] = {
                        "unit": unit_val,
                        "prompt": prompt_val,
                        "value": final_val,
                        "row": r,
                        "col": c,
                        "col_letter": get_column_letter(c),
                    }
                else:
                    per_project[c]["opex_year1"][current_category][key] = final_val
            continue

        # 일반 필드 저장
        for c in project_col_indices:
            v = row_vals[c]
            if include_meta:
                per_project[c]["opex_year1"][current_category][key] = {
                    "unit": unit_val,
                    "prompt": prompt_val,
                    "value": v,
                    "row": r,
                    "col": c,
                    "col_letter": get_column_letter(c),
                }
            else:
                per_project[c]["opex_year1"][current_category][key] = v

    projects = [per_project[c] for c in project_col_indices]

    return {
        "source": {
            "file": xlsx_path,
            "sheet": ws.title,
            "title_row": title_row,
            "label_col": get_column_letter(label_col),
            "unit_col": get_column_letter(unit_col),
            "projects_detected": [project_cols_map[c] for c in project_col_indices],
            "project_columns": project_columns_list,
        },
        "projects": projects,
    }


def save_json(out_name: str, obj: Dict[str, Any]) -> Path:
    out_arg = Path(out_name)
    out_path = out_arg if out_arg.is_absolute() else (Path.cwd() / out_arg)
    if len(out_path.parts) == 1:
        out_path = Path.cwd() / "data" / out_path.name
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2, default=str)
    return out_path


def main():
    ap = argparse.ArgumentParser(description="Extract PV solar inputs from Excel into JSON.")
    ap.add_argument("xlsx", help="Input xlsx path")
    ap.add_argument("-o", "--out", default="result.json", help="Output JSON filename (saved under ./data/)")
    ap.add_argument(
        "--mode",
        choices=["gi_prod", "opex_year1", "all"],
        default="gi_prod",
        help="gi_prod: active sheet General inputs+Production, opex_year1: Opex - year 1 sheet, all: both",
    )
    ap.add_argument(
        "-c",
        "--label-col",
        default="B",
        help="(gi_prod only) Label column letter for General inputs/Production (default: B)",
    )
    ap.add_argument("--include-meta", action="store_true", help="Include row/col/unit/prompt metadata")
    args = ap.parse_args()

    if args.mode == "gi_prod":
        result = extract_gi_and_production_from_active_sheet(
            xlsx_path=args.xlsx,
            label_col_letter=args.label_col,
            include_meta=args.include_meta,
        )
        out_path = save_json(args.out, result)
        print(f"[OK] saved: {out_path}")
        print("     mode: gi_prod")
        print(f"     sheet: {result['source']['sheet']}")
        print(f"     projects: {len(result['projects'])}")

    elif args.mode == "opex_year1":
        result = extract_opex_year1_from_sheet(
            xlsx_path=args.xlsx,
            include_meta=args.include_meta,
        )
        out_path = save_json(args.out, result)
        print(f"[OK] saved: {out_path}")
        print("     mode: opex_year1")
        print(f"     sheet: {result['source']['sheet']}")
        print(f"     projects: {len(result['projects'])}")

    else:  # all
        gi_prod = extract_gi_and_production_from_active_sheet(
            xlsx_path=args.xlsx,
            label_col_letter=args.label_col,
            include_meta=args.include_meta,
        )
        opex1 = extract_opex_year1_from_sheet(
            xlsx_path=args.xlsx,
            include_meta=args.include_meta,
        )
        out1 = save_json("gi_prod.json", gi_prod)
        out2 = save_json("opex_year1.json", opex1)
        print(f"[OK] saved: {out1}")
        print(f"[OK] saved: {out2}")
        print("     mode: all")
        print(f"     gi_prod sheet: {gi_prod['source']['sheet']} / projects: {len(gi_prod['projects'])}")
        print(f"     opex_year1 sheet: {opex1['source']['sheet']} / projects: {len(opex1['projects'])}")


if __name__ == "__main__":
    main()
"""

python ./pv_solar/test3/extract_general_inputs.py "/path/to/test2.xlsx" --mode all

"""
