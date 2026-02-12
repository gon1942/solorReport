#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import re
import datetime
from pathlib import Path
from typing import Any, Dict, Tuple, Optional, List

import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string

PROJECT_RE = re.compile(r"^Project\s*\d+\s*$", re.IGNORECASE)


def to_json_value(v: Any) -> Any:
    """openpyxl 값 -> JSON 직렬화 가능한 값으로 변환"""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def find_project_header_row(ws, scan_max_rows: int = 300) -> Tuple[int, Dict[int, str]]:
    """
    'Project 1'...'Project n' 헤더가 있는 행과,
    프로젝트 열 인덱스 -> 헤더명 매핑 반환
    """
    best_row: Optional[int] = None
    best_map: Dict[int, str] = {}

    max_rows = min(ws.max_row, scan_max_rows)
    for r in range(1, max_rows + 1):
        row_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and PROJECT_RE.match(v.strip()):
                row_map[c] = v.strip()
        if row_map:
            if best_row is None or len(row_map) > len(best_map):
                best_row = r
                best_map = row_map

    if best_row is None:
        raise RuntimeError("Project 헤더 행(Project 1, Project 2, ...)을 찾지 못했습니다.")

    # 연속 블록만 유지(오탐 방지)
    cols = sorted(best_map.keys())
    contiguous = [cols[0]]
    for i in range(1, len(cols)):
        if cols[i] == cols[i - 1] + 1:
            contiguous.append(cols[i])
        else:
            break

    project_cols = {c: best_map[c] for c in contiguous}
    return best_row, project_cols


def is_section_header(ws, r: int, label_col: int, probe_value_col: int) -> bool:
    """
    섹션 헤더 행 판별(프로브 컬럼 기준):
    - label_col에 문자열이 있고
    - probe_value_col은 비어있고
    - label 오른쪽(타입/설명)도 비어있는 경우가 많음
    """
    label = ws.cell(r, label_col).value
    if not isinstance(label, str) or not label.strip():
        return False

    if ws.cell(r, probe_value_col).value is not None:
        return False

    right = ws.cell(r, label_col + 1).value
    return right is None


def find_section_block(
    ws,
    label_col: int,
    probe_value_col: int,
    section_titles: List[str],
) -> Optional[Tuple[int, int]]:
    """
    섹션(예: General inputs, Production) 범위를 찾는다.
    - section_titles: 허용 헤더 문자열 목록(소문자 비교)

    반환: (start_row_inclusive, end_row_exclusive) 또는 None(섹션 미발견)
    """
    titles = {t.strip().lower() for t in section_titles if t and t.strip()}
    start_header: Optional[int] = None

    # 섹션 시작 찾기
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, label_col).value
        if isinstance(v, str) and v.strip().lower() in titles:
            start_header = r
            break

    if start_header is None:
        return None

    # 섹션 끝 찾기(다음 섹션 헤더가 나오기 전)
    end = ws.max_row + 1
    for r in range(start_header + 1, ws.max_row + 1):
        if is_section_header(ws, r, label_col, probe_value_col):
            v = ws.cell(r, label_col).value
            if isinstance(v, str) and v.strip().lower() not in titles:
                end = r
                break

    return start_header + 1, end


def extract_section_kv_for_project(
    ws,
    start_row: int,
    end_row: int,
    label_col: int,
    type_col: int,
    value_col: int,
    probe_value_col: int,
    include_meta: bool,
) -> Dict[str, Any]:
    """
    지정된 섹션 블록(start_row~end_row)에서
    label_col의 키와 value_col의 값을 추출하여 dict 반환
    """
    out: Dict[str, Any] = {}

    for r in range(start_row, end_row):
        label = ws.cell(r, label_col).value
        if not isinstance(label, str) or not label.strip():
            continue

        # 안전장치: 섹션 끝(다음 섹션 헤더로 판단)이면 종료
        if is_section_header(ws, r, label_col, probe_value_col):
            # 블록 end_row로 이미 잘렸어야 하지만, 문서 변형 대비
            break

        key = label.strip()
        ftype = ws.cell(r, type_col).value
        val = to_json_value(ws.cell(r, value_col).value)

        if include_meta:
            out[key] = {
                "type": ftype,
                "value": val,
                "row": r,
                "col": value_col,
                "col_letter": get_column_letter(value_col),
            }
        else:
            out[key] = val

    return out


def extract_all_projects_general_inputs_and_production(
    xlsx_path: str,
    label_col_letter: str = "B",
    include_meta: bool = False,
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active  # 활성 시트만

    header_row, project_cols_map = find_project_header_row(ws)

    label_col = column_index_from_string(label_col_letter)
    type_col = label_col + 1

    project_col_indices = sorted(project_cols_map.keys())
    probe_col = project_col_indices[0]  # 섹션 헤더 판별용(프로브)

    # 섹션 범위 찾기
    gi_block = find_section_block(ws, label_col, probe_col, ["General inputs"])
    prod_block = find_section_block(ws, label_col, probe_col, ["Production"])

    missing_sections = []
    if gi_block is None:
        missing_sections.append("General inputs")
    if prod_block is None:
        missing_sections.append("Production")

    projects: List[Dict[str, Any]] = []

    for col_idx in project_col_indices:
        pname = project_cols_map[col_idx]
        col_letter = get_column_letter(col_idx)

        general_inputs = {}
        production = {}

        if gi_block is not None:
            gi_start, gi_end = gi_block
            general_inputs = extract_section_kv_for_project(
                ws=ws,
                start_row=gi_start,
                end_row=gi_end,
                label_col=label_col,
                type_col=type_col,
                value_col=col_idx,
                probe_value_col=probe_col,
                include_meta=include_meta,
            )

        if prod_block is not None:
            pr_start, pr_end = prod_block
            production = extract_section_kv_for_project(
                ws=ws,
                start_row=pr_start,
                end_row=pr_end,
                label_col=label_col,
                type_col=type_col,
                value_col=col_idx,
                probe_value_col=probe_col,
                include_meta=include_meta,
            )

        projects.append(
            {
                "name": pname,
                "col": col_letter,
                "general_inputs": general_inputs,
                "production": production,
            }
        )

    return {
        "source": {
            "file": xlsx_path,
            "sheet": ws.title,
            "project_header_row": header_row,
            "label_col": label_col_letter,
            "projects_detected": [project_cols_map[c] for c in project_col_indices],
            "project_columns": [{"project": project_cols_map[c], "col": get_column_letter(c)} for c in project_col_indices],
            "missing_sections": missing_sections,
        },
        "projects": projects,
    }


def main():
    ap = argparse.ArgumentParser(
        description="Extract 'General inputs' + 'Production' for ALL projects from the active sheet into JSON."
    )
    ap.add_argument("xlsx", help="Input xlsx path")
    ap.add_argument("-c", "--label-col", default="B", help="Label column letter (default: B)")
    ap.add_argument("-o", "--out", default="projects_general_inputs_production.json", help="Output JSON filename")
    ap.add_argument("--include-meta", action="store_true", help="Include type/row/col metadata")
    ap.add_argument("--indent", type=int, default=2, help="JSON indent (default: 2)")
    args = ap.parse_args()

    result = extract_all_projects_general_inputs_and_production(
        xlsx_path=args.xlsx,
        label_col_letter=args.label_col,
        include_meta=args.include_meta,
    )

    out_dir = Path.cwd() / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / Path(args.out).name

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=args.indent)

    print(f"[OK] saved: {out_path}")
    print(f"     sheet: {result['source']['sheet']}")
    print(f"     projects: {len(result['projects'])}")
    if result["source"].get("missing_sections"):
        print(f"     missing_sections: {result['source']['missing_sections']}")


if __name__ == "__main__":
    main()
