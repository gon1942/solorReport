#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import re
import datetime
import sys
from pathlib import Path
from typing import Any, Dict, Tuple, Optional, List

import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string

PROJECT_RE = re.compile(r"^Project\s*\d+\s*$", re.IGNORECASE)

CAPEX_FIELD_LABELS = {
    "currency",
    "cost",
    "indexed?",
    "indexation reference date",
    "indexation %",
    "monthly distribution",
    "comments",
    "unit",
}

EXCEL_ZERO_DATE = datetime.date(1899, 12, 30)


# -----------------------------
# Utils
# -----------------------------

def norm_text(v: Any) -> Optional[str]:
    if not isinstance(v, str):
        return None
    s = v.strip()
    return s if s else None


def to_json_value(v: Any) -> Any:
    if v is None:
        return None

    if isinstance(v, datetime.datetime):
        return v.isoformat()

    if isinstance(v, datetime.date):
        if v == EXCEL_ZERO_DATE:
            return ""
        return v.isoformat()

    if isinstance(v, datetime.time):
        return v.isoformat()

    if isinstance(v, float) and v.is_integer():
        return int(v)

    if isinstance(v, str):
        return v.strip()

    return v


def blank_if_empty(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, str):
        s = v.strip()
        if s == "-" or s == "":
            return ""
        return s
    return v


def normalize_indexed_value(v: Any) -> str:
    """
    Indexed? 값 정규화:
    - 1 / "1" / True  -> "Yes"
    - 0 / "0" / False -> "No"
    - "Yes"/"No" 형태면 그대로
    - 빈값이면 ""
    """
    v = blank_if_empty(v)
    if v == "":
        return ""

    if isinstance(v, str):
        s = v.strip().lower()
        if s in ("yes", "y", "true"):
            return "Yes"
        if s in ("no", "n", "false"):
            return "No"
        if s == "1":
            return "Yes"
        if s == "0":
            return "No"
        return v.strip()

    if isinstance(v, bool):
        return "Yes" if v else "No"

    if isinstance(v, (int, float)):
        try:
            return "Yes" if int(v) == 1 else "No"
        except Exception:
            return str(v)

    return str(v)


def is_column_visible(ws, col_idx: int) -> bool:
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    if dim is None:
        return True

    if getattr(dim, "hidden", False):
        return False

    width = getattr(dim, "width", None)
    if width is not None:
        try:
            if float(width) == 0.0:
                return False
        except Exception:
            pass

    return True


# -----------------------------
# Project header detection (any sheet)
# -----------------------------

def find_project_header_row_visible(
    ws,
    scan_max_rows: int = 1500,
) -> Tuple[int, Dict[int, str]]:
    """
    ws 내에서 Project 1..n이 존재하는 행을 찾아,
    '보이는 열'만 col_idx -> "Project n" 형태로 반환.
    """
    best_row: Optional[int] = None
    best_visible: Dict[int, str] = {}

    max_rows = min(ws.max_row, scan_max_rows)
    for r in range(1, max_rows + 1):
        row_visible: Dict[int, str] = {}

        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and PROJECT_RE.match(v.strip()):
                if is_column_visible(ws, c):
                    row_visible[c] = v.strip()

        if row_visible:
            if best_row is None or len(row_visible) > len(best_visible):
                best_row = r
                best_visible = row_visible

    if best_row is None:
        raise RuntimeError("Project 헤더 행(Project 1, Project 2, ...)을 찾지 못했습니다.")

    # contiguous 구간만 사용
    cols = sorted(best_visible.keys())
    contiguous = [cols[0]]
    for i in range(1, len(cols)):
        if cols[i] == cols[i - 1] + 1:
            contiguous.append(cols[i])
        else:
            break

    # 중복 Project명 제거
    seen = set()
    dedup: Dict[int, str] = {}
    for c in contiguous:
        name = best_visible[c]
        if name in seen:
            continue
        seen.add(name)
        dedup[c] = name

    return best_row, dedup


def find_best_project_map_in_workbook(wb) -> Optional[Dict[str, str]]:
    """
    워크북 전체에서 Project 1..n 헤더가 가장 잘 잡히는 시트를 찾아
    col_letter -> project_name 형태로 반환.
    """
    best_map: Optional[Dict[str, str]] = None
    best_cnt = 0

    for ws in wb.worksheets:
        try:
            _, mp = find_project_header_row_visible(ws, scan_max_rows=1500)
            if mp and len(mp) > best_cnt:
                best_cnt = len(mp)
                best_map = {get_column_letter(c): mp[c] for c in sorted(mp.keys())}
        except Exception:
            continue

    return best_map


# -----------------------------
# Capex sheet locate
# -----------------------------

def find_capex_sheet_or_ws(wb) -> Any:
    """
    1) Unit + Cost... 형태의 CAPEX 테이블이 실제로 탐지되는 시트를 우선 선택
       - INPUTSHEET 계열 우선
       - CAPEX 시트명 계열 차순위
    2) 그래도 없으면 기존 규칙(capex 시트명/셀 텍스트) 사용
    3) 최종적으로 active 반환
    """
    prioritized: list[Any] = []
    secondary: list[Any] = []
    others: list[Any] = []

    for ws in wb.worksheets:
        name = (ws.title or "").lower()
        if "inputsheet" in name:
            prioritized.append(ws)
        elif "capex" in name:
            secondary.append(ws)
        else:
            others.append(ws)

    for ws in prioritized + secondary + others:
        try:
            find_capex_table_header(ws)
            return ws
        except Exception:
            continue

    for ws in secondary:
        return ws

    target = "capex"
    for ws in wb.worksheets:
        max_rows = min(ws.max_row, 800)
        max_cols = min(ws.max_column, 200)
        for r in range(1, max_rows + 1):
            for c in range(1, max_cols + 1):
                v = norm_text(ws.cell(r, c).value)
                if v and v.strip().lower() == target:
                    return ws

    return wb.active


# -----------------------------
# Capex table header & label_col auto-detect
# -----------------------------

def find_capex_table_header(ws, scan_max_rows: int = 250, scan_max_cols: int = 200) -> Tuple[int, int]:
    """
    'Unit' 셀이 있는 행 중, 오른쪽에 'Cost'가 여러 개 있는 행을 Capex 표 헤더로 추정.
    반환: (header_row, unit_col)
    """
    best = None  # (cost_count, row, col)
    max_rows = min(ws.max_row, scan_max_rows)
    max_cols = min(ws.max_column, scan_max_cols)

    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            v = norm_text(ws.cell(r, c).value)
            if not v or v.strip().lower() != "unit":
                continue

            # 오른쪽에 Cost 몇 개나 있는지 점수화
            cost_count = 0
            for cc in range(c + 1, min(max_cols, c + 25) + 1):
                vv = norm_text(ws.cell(r, cc).value)
                if vv and vv.strip().lower() == "cost":
                    cost_count += 1

            if cost_count >= 2:
                cand = (cost_count, r, c)
                if best is None or cand[0] > best[0]:
                    best = cand

    if best is None:
        raise RuntimeError("Capex 표 헤더( Unit + Cost... )를 찾지 못했습니다. 헤더 구조를 확인해 주십시오.")

    _, header_row, unit_col = best
    return header_row, unit_col


def detect_label_col(ws, header_row: int, unit_col: int) -> int:
    """
    header_row 아래 구간에서 CAPEX_FIELD_LABELS가 가장 많이 등장하는 컬럼을 label_col로 선택.
    (Capex 시트에서 A가 category, B가 필드라벨인 케이스 해결)
    """
    max_rows = min(ws.max_row, header_row + 350)
    candidates = list(range(1, max(1, unit_col)))  # 1..unit_col-1
    if not candidates:
        return 1

    best_col = candidates[0]
    best_score = -1

    for col in candidates:
        score = 0
        for r in range(header_row + 1, max_rows + 1):
            v = norm_text(ws.cell(r, col).value)
            if not v:
                continue
            lw = v.strip().lower()

            # 필드 라벨 직접 매치: 높은 점수
            if lw in CAPEX_FIELD_LABELS:
                score += 5
                continue

            # 카테고리 헤더처럼 보이고, 아래 5줄 내 필드라벨이 나오면 가산
            for rr in range(r + 1, min(r + 6, max_rows + 1)):
                nxt = norm_text(ws.cell(rr, col).value)
                if nxt and nxt.strip().lower() in CAPEX_FIELD_LABELS:
                    score += 1
                    break

        if score > best_score:
            best_score = score
            best_col = col

    return best_col


def choose_prompt_col(label_col: int, unit_col: int) -> int:
    """
    prompt_col은 보통 unit_col-1 위치. 단 label_col과 겹치면 label_col+1로 조정.
    """
    if unit_col <= 1:
        return max(1, label_col + 1)

    pc = unit_col - 1
    if pc == label_col:
        if label_col + 1 < unit_col:
            pc = label_col + 1
    return pc


# -----------------------------
# Project cols robust builder
# -----------------------------

def column_has_data(ws, col: int, rows: List[int]) -> bool:
    for r in rows:
        if r < 1 or r > ws.max_row:
            continue
        v = ws.cell(r, col).value
        vv = blank_if_empty(to_json_value(v))
        if vv != "":
            return True
    return False


def build_project_cols_for_table_robust(
    ws,
    header_row: int,
    unit_col: int,
    global_col_to_project: Optional[Dict[str, str]] = None,
) -> Tuple[Dict[int, str], List[Dict[str, str]]]:
    """
    프로젝트 열 매핑:
    1) 현재 ws에 Project 1..n 헤더 있으면 사용
    2) global_col_to_project로 (열 문자 기준) 매핑
    3) 헤더 없이도 실제 데이터가 있는 열을 기준으로 Project 1..n 생성
    """
    # 1) 현재 시트에서 Project 헤더 찾기
    try:
        _, mp = find_project_header_row_visible(ws, scan_max_rows=1500)
        if mp:
            cols = sorted(mp.keys())
            return mp, [{"project": mp[c], "col": get_column_letter(c)} for c in cols]
    except Exception:
        pass

    # 2) global map 적용
    project_cols_map: Dict[int, str] = {}
    if global_col_to_project:
        for col_letter, pname in global_col_to_project.items():
            col_idx = column_index_from_string(col_letter)
            if col_idx <= ws.max_column and col_idx > unit_col and is_column_visible(ws, col_idx):
                project_cols_map[col_idx] = pname

        if project_cols_map:
            cols = sorted(project_cols_map.keys())
            return project_cols_map, [{"project": project_cols_map[c], "col": get_column_letter(c)} for c in cols]

    # 3) 데이터 기반 탐지 (header_row 주변)
    sample_rows = [header_row, header_row + 1, header_row + 2, header_row + 3, header_row + 4]

    started = False
    empty_run = 0
    max_empty_run_after_start = 6
    idx = 1

    for c in range(unit_col + 1, ws.max_column + 1):
        if not is_column_visible(ws, c):
            continue

        has = column_has_data(ws, c, sample_rows)

        if has:
            started = True
            empty_run = 0
            col_letter = get_column_letter(c)
            pname = global_col_to_project.get(col_letter) if global_col_to_project else None
            if not pname:
                pname = f"Project {idx}"
            project_cols_map[c] = pname
            idx += 1
        else:
            if started:
                empty_run += 1
                if empty_run >= max_empty_run_after_start:
                    break

    # 중복 제거
    seen = set()
    dedup: Dict[int, str] = {}
    for c in sorted(project_cols_map.keys()):
        n = project_cols_map[c]
        if n in seen:
            continue
        seen.add(n)
        dedup[c] = n

    project_cols_map = dedup
    cols = sorted(project_cols_map.keys())
    return project_cols_map, [{"project": project_cols_map[c], "col": get_column_letter(c)} for c in cols]


# -----------------------------
# Capex parsing helpers
# -----------------------------

def is_category_header_generic(ws, r: int, label_col: int, field_labels: set) -> bool:
    label = norm_text(ws.cell(r, label_col).value)
    if not label:
        return False
    if label.strip().lower() in field_labels:
        return False

    for rr in range(r + 1, min(r + 6, ws.max_row + 1)):
        nxt = norm_text(ws.cell(rr, label_col).value)
        if nxt and nxt.strip().lower() in field_labels:
            return True
    return False


def extract_month_id_from_row(ws, r: int, label_col: int, prompt_col: int, unit_col: int) -> Optional[int]:
    candidates = [
        ws.cell(r, unit_col).value,
        ws.cell(r, prompt_col).value,
        ws.cell(r, label_col).value,
    ]
    for v in candidates:
        vv = to_json_value(v)
        if vv is None:
            continue
        if isinstance(vv, int) and 1 <= vv <= 24:
            return vv
        if isinstance(vv, float):
            try:
                iv = int(vv)
                if iv == vv and 1 <= iv <= 24:
                    return iv
            except Exception:
                pass
        if isinstance(vv, str):
            s = vv.strip()
            if s.isdigit():
                iv = int(s)
                if 1 <= iv <= 24:
                    return iv
    return None


# -----------------------------
# Main extractor
# -----------------------------

def extract_capex_from_workbook(xlsx_path: str, include_meta: bool = False) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    global_map = find_best_project_map_in_workbook(wb)

    ws = find_capex_sheet_or_ws(wb)

    # 1) Capex 표 헤더( Unit + Cost... ) 찾기
    header_row, unit_col = find_capex_table_header(ws)

    # 2) label_col 자동 탐지 (A/B 뒤바뀜 대응)
    label_col = detect_label_col(ws, header_row, unit_col)
    prompt_col = choose_prompt_col(label_col, unit_col)

    # 3) 프로젝트 열 매핑
    project_cols_map, project_columns_list = build_project_cols_for_table_robust(
        ws=ws,
        header_row=header_row,
        unit_col=unit_col,
        global_col_to_project=global_map,
    )
    project_col_indices = sorted(project_cols_map.keys())
    if not project_col_indices:
        raise RuntimeError("Capex에서 프로젝트 값 열을 찾지 못했습니다. Unit 오른쪽 값 열에 데이터가 있는지 확인해 주십시오.")

    per_project: Dict[int, Dict[str, Any]] = {}
    for c in project_col_indices:
        per_project[c] = {
            "name": project_cols_map[c],
            "col": get_column_letter(c),
            "capex": {},
        }

    current_category: Optional[str] = None
    in_monthly = False

    for r in range(header_row + 1, ws.max_row + 1):
        label = norm_text(ws.cell(r, label_col).value)

        if not label and not in_monthly:
            continue

        # 카테고리 헤더
        if label and is_category_header_generic(ws, r, label_col, CAPEX_FIELD_LABELS):
            current_category = label
            in_monthly = False
            for c in project_col_indices:
                per_project[c]["capex"].setdefault(current_category, {})

            # enabled (카테고리 헤더행 값)
            any_val = False
            header_vals: Dict[int, Any] = {}
            for c in project_col_indices:
                v = blank_if_empty(to_json_value(ws.cell(r, c).value))
                header_vals[c] = v
                if v != "":
                    any_val = True
            if any_val:
                for c in project_col_indices:
                    per_project[c]["capex"][current_category]["enabled"] = header_vals[c]
            continue

        if current_category is None:
            current_category = "Capex"
            for c in project_col_indices:
                per_project[c]["capex"].setdefault(current_category, {})

        # monthly distribution
        if label and label.strip().lower() == "monthly distribution":
            in_monthly = True
            for c in project_col_indices:
                per_project[c]["capex"][current_category].setdefault("Monthly distribution", {})
        else:
            if in_monthly and label:
                lwr = label.strip().lower()
                if lwr in CAPEX_FIELD_LABELS and lwr != "monthly distribution":
                    in_monthly = False

        unit_val = blank_if_empty(to_json_value(ws.cell(r, unit_col).value))
        prompt_val = blank_if_empty(to_json_value(ws.cell(r, prompt_col).value))

        if in_monthly:
            month_id = extract_month_id_from_row(ws, r, label_col, prompt_col, unit_col)
            if month_id is None:
                continue

            for c in project_col_indices:
                v = blank_if_empty(to_json_value(ws.cell(r, c).value))
                if include_meta:
                    per_project[c]["capex"][current_category]["Monthly distribution"][str(month_id)] = {
                        "unit": unit_val,
                        "prompt": prompt_val,
                        "value": v,
                        "row": r,
                        "col": c,
                        "col_letter": get_column_letter(c),
                    }
                else:
                    per_project[c]["capex"][current_category]["Monthly distribution"][str(month_id)] = v
            continue

        if not label:
            continue

        key = label
        key_lower = key.strip().lower()

        row_vals: Dict[int, Any] = {}
        for c in project_col_indices:
            v = blank_if_empty(to_json_value(ws.cell(r, c).value))
            if key_lower == "indexed?":
                v = normalize_indexed_value(v)
            row_vals[c] = v

        # Comments: 값이 전부 비면 prompt(머지셀 설명)를 복제
        if key_lower == "comments":
            any_project_comment = any(row_vals[c] != "" for c in project_col_indices)
            final_comment_prompt = prompt_val if (not any_project_comment and prompt_val != "") else ""
            for c in project_col_indices:
                final_val = row_vals[c] if any_project_comment else (final_comment_prompt or "")
                if include_meta:
                    per_project[c]["capex"][current_category][key] = {
                        "unit": unit_val,
                        "prompt": prompt_val,
                        "value": final_val,
                        "row": r,
                        "col": c,
                        "col_letter": get_column_letter(c),
                    }
                else:
                    per_project[c]["capex"][current_category][key] = final_val
            continue

        for c in project_col_indices:
            v = row_vals[c]
            if include_meta:
                per_project[c]["capex"][current_category][key] = {
                    "unit": unit_val,
                    "prompt": prompt_val,
                    "value": v,
                    "row": r,
                    "col": c,
                    "col_letter": get_column_letter(c),
                }
            else:
                per_project[c]["capex"][current_category][key] = v

    projects = [per_project[c] for c in project_col_indices]

    return {
        "source": {
            "file": xlsx_path,
            "sheet": ws.title,
            "header_row": header_row,
            "label_col": get_column_letter(label_col),
            "prompt_col": get_column_letter(prompt_col),
            "unit_col": get_column_letter(unit_col),
            "projects_detected": [project_cols_map[c] for c in project_col_indices],
            "project_columns": project_columns_list,
            "used_global_project_map": bool(global_map),
        },
        "projects": projects,
    }


def save_json(out_name: str, obj: Dict[str, Any], indent: int) -> Path:
    out_arg = Path(out_name)
    out_path = out_arg if out_arg.is_absolute() else (Path.cwd() / out_arg)
    if len(out_path.parts) == 1:
        out_path = Path.cwd() / "data" / out_path.name
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=indent, default=str)
    return out_path


def main():
    ap = argparse.ArgumentParser(description="Extract Capex from Excel into JSON.")
    ap.add_argument("xlsx", help="Input xlsx path")
    ap.add_argument("--mode", choices=["capex"], default="capex", help="capex only")
    ap.add_argument("-o", "--out", default="capex.json", help="Output JSON filename (saved under ./data/)")
    ap.add_argument("--indent", type=int, default=2, help="JSON indent (default: 2)")
    ap.add_argument("--include-meta", action="store_true", help="Include metadata (row/col/unit/prompt)")
    args = ap.parse_args()

    try:
        result = extract_capex_from_workbook(args.xlsx, include_meta=args.include_meta)
        out_path = save_json(args.out, result, indent=args.indent)
        print(f"[OK] saved: {out_path}")
        print(f"     sheet: {result['source']['sheet']}")
        print(f"     header_row: {result['source']['header_row']}")
        print(f"     label_col: {result['source']['label_col']}, prompt_col: {result['source']['prompt_col']}, unit_col: {result['source']['unit_col']}")
        print(f"     projects: {len(result['projects'])}")
    except Exception as e:
        # apport(excepthook)로 2차 에러나는 환경이 있어 여기서 종료
        print(f"[ERROR] {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
