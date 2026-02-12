#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import re
import datetime
from pathlib import Path
from typing import Any, Dict, Tuple, Optional, List

import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string

PROJECT_RE = re.compile(r"^Project\s*\d+\s*$", re.IGNORECASE)

# 문서에 존재할 수 있는 "Top-level 섹션명" 목록 (필요 시 추가)
TOP_LEVEL_SECTIONS = [
    "General inputs",
    "Production",
    "CAPEX",
    "OPEX",
    "Revenues",
    "Financing",
    "Debt",
    "Equity",
    "Assumptions",
]


def norm_text(v: Any) -> Optional[str]:
    if not isinstance(v, str):
        return None
    s = v.strip()
    return s if s else None


# def to_json_value(v: Any) -> Any:
#     """
#     openpyxl 값 -> JSON 직렬화 가능한 값으로 변환
#     - date/datetime/time 모두 ISO 문자열로 변환
#     """
#     if v is None:
#         return None

#     if isinstance(v, (datetime.datetime, datetime.date)):
#         return v.isoformat()

#     # ✅ 추가: time 타입 처리
#     if isinstance(v, datetime.time):
#         # "HH:MM:SS" 형태
#         return v.isoformat()

#     # Excel의 정수처럼 보이는 float는 int로 정규화
#     if isinstance(v, float) and v.is_integer():
#         return int(v)

#     # 문자열은 trim
#     if isinstance(v, str):
#         return v.strip()

#     return v
def to_json_value(v: Any) -> Any:
    if v is None:
        return None

    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()

    # ✅ time 타입도 JSON 가능하게
    if isinstance(v, datetime.time):
        return v.isoformat()

    if isinstance(v, float) and v.is_integer():
        return int(v)

    if isinstance(v, str):
        return v.strip()

    return v


def is_column_visible(ws, col_idx: int) -> bool:
    """
    '실질적으로 눈에 보이는 열' 판별
    - hidden=True 또는 width=0이면 제외
    """
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    if dim is None:
        return True

    if getattr(dim, "hidden", False):
        return False

    width = getattr(dim, "width", None)
    if width is not None:
        try:
            if float(width) == 0.0:
                return False
        except Exception:
            pass

    return True


def find_project_header_row_visible(
    ws,
    scan_max_rows: int = 300,
) -> Tuple[int, Dict[int, str], List[Dict[str, str]], List[Dict[str, str]]]:
    """
    - Project 헤더 행을 찾고
    - '보이는 열'의 Project 컬럼만 반환
    - 같은 Project명이 중복이면 뒤쪽 제거
    반환:
      (header_row,
       project_cols_map: {col_idx: "Project n"},
       hidden_project_cols: [{"project":..., "col":...}],
       dropped_duplicates: [{"project":..., "col":..., "reason":...}])
    """
    best_row: Optional[int] = None
    best_visible: Dict[int, str] = {}
    best_hidden: Dict[int, str] = {}

    max_rows = min(ws.max_row, scan_max_rows)
    for r in range(1, max_rows + 1):
        row_visible: Dict[int, str] = {}
        row_hidden: Dict[int, str] = {}

        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and PROJECT_RE.match(v.strip()):
                if is_column_visible(ws, c):
                    row_visible[c] = v.strip()
                else:
                    row_hidden[c] = v.strip()

        if row_visible:
            if best_row is None or len(row_visible) > len(best_visible):
                best_row = r
                best_visible = row_visible
                best_hidden = row_hidden

    if best_row is None:
        raise RuntimeError("Project 헤더 행(Project 1, Project 2, ...)을 찾지 못했습니다.")

    # 연속 블록만 유지(오탐 방지)
    cols = sorted(best_visible.keys())
    contiguous = [cols[0]]
    for i in range(1, len(cols)):
        if cols[i] == cols[i - 1] + 1:
            contiguous.append(cols[i])
        else:
            break

    # 중복 Project명 제거(뒤쪽 제거)
    seen = set()
    dedup_map: Dict[int, str] = {}
    dropped_duplicates: List[Dict[str, str]] = []
    for c in contiguous:
        name = best_visible[c]
        if name in seen:
            dropped_duplicates.append(
                {"project": name, "col": get_column_letter(c), "reason": "duplicate_project_name"}
            )
            continue
        seen.add(name)
        dedup_map[c] = name

    hidden_project_cols = [
        {"project": best_hidden[c], "col": get_column_letter(c)}
        for c in sorted(best_hidden.keys())
    ]

    return best_row, dedup_map, hidden_project_cols, dropped_duplicates


def find_section_block_by_top_sections(
    ws,
    label_col: int,
    section_name: str,
    top_sections: List[str],
) -> Optional[Tuple[int, int]]:
    """
    섹션 범위를 '다음 Top-level 섹션명'이 나올 때까지로 계산합니다.
    (Production 내부 소제목은 종료 조건이 되면 안 되므로 이 방식이 필요합니다.)
    """
    target = section_name.strip().lower()
    top_set = {s.strip().lower() for s in top_sections if s and s.strip()}

    start_header: Optional[int] = None
    for r in range(1, ws.max_row + 1):
        v = norm_text(ws.cell(r, label_col).value)
        if v and v.lower() == target:
            start_header = r
            break

    if start_header is None:
        return None

    end = ws.max_row + 1
    for r in range(start_header + 1, ws.max_row + 1):
        v = norm_text(ws.cell(r, label_col).value)
        if v and v.lower() in top_set and v.lower() != target:
            end = r
            break

    return start_header + 1, end


def is_group_header_row(ws, r: int, label_col: int, type_col: int, value_col: int) -> bool:
    """
    Production 내부 소제목/그룹 헤더 판별:
    - label 있음
    - type_col 비어있음
    - value_col 비어있음
    """
    label = norm_text(ws.cell(r, label_col).value)
    if not label:
        return False
    t = ws.cell(r, type_col).value
    if t is not None and norm_text(t):
        return False
    if ws.cell(r, value_col).value is not None:
        return False
    return True


def extract_section_kv_for_project_flat(
    ws,
    start_row: int,
    end_row: int,
    label_col: int,
    type_col: int,
    value_col: int,
    include_meta: bool,
    include_empty: bool,
) -> Dict[str, Any]:
    """
    General inputs처럼 단순 Key-Value로 추출
    """
    out: Dict[str, Any] = {}
    for r in range(start_row, end_row):
        key = norm_text(ws.cell(r, label_col).value)
        if not key:
            continue

        val = to_json_value(ws.cell(r, value_col).value)
        if (val is None) and (not include_empty):
            continue

        if include_meta:
            out[key] = {
                "type": to_json_value(ws.cell(r, type_col).value),
                "value": val,
                "row": r,
                "col": value_col,
                "col_letter": get_column_letter(value_col),
            }
        else:
            out[key] = val
    return out


def extract_section_kv_for_project_grouped(
    ws,
    start_row: int,
    end_row: int,
    label_col: int,
    type_col: int,
    value_col: int,
    include_meta: bool,
    include_empty: bool,
) -> Dict[str, Any]:
    """
    Production처럼 소제목(그룹)이 있는 섹션을 그룹화하여 추출
    - include_empty=True이면 빈 값도 포함하고 value는 ""(공백 문자열)로 저장
    """
    out: Dict[str, Any] = {}
    current_group: Optional[str] = None

    for r in range(start_row, end_row):
        label = norm_text(ws.cell(r, label_col).value)
        if not label:
            continue

        # 그룹 헤더(소제목)
        if is_group_header_row(ws, r, label_col, type_col, value_col):
            current_group = label
            # 그룹이 미리 생성되길 원하면 아래 한 줄 활성화
            # out.setdefault(current_group, {})
            continue

        raw_val = ws.cell(r, value_col).value
        val = to_json_value(raw_val)

        # ✅ 빈값 처리: include_empty=True면 ""로 저장
        if val is None:
            if not include_empty:
                continue
            val = ""

        if current_group:
            grp = out.setdefault(current_group, {})
            if include_meta:
                raw_type = ws.cell(r, type_col).value
                tval = to_json_value(raw_type)
                if tval is None and include_empty:
                    tval = ""
                grp[label] = {
                    "type": tval,
                    "value": val,
                    "row": r,
                    "col": value_col,
                    "col_letter": get_column_letter(value_col),
                }
            else:
                grp[label] = val
        else:
            if include_meta:
                raw_type = ws.cell(r, type_col).value
                tval = to_json_value(raw_type)
                if tval is None and include_empty:
                    tval = ""
                out[label] = {
                    "type": tval,
                    "value": val,
                    "row": r,
                    "col": value_col,
                    "col_letter": get_column_letter(value_col),
                }
            else:
                out[label] = val

    return out


def extract_all_projects_gi_and_production_visible(
    xlsx_path: str,
    label_col_letter: str = "B",
    include_meta: bool = False,
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active  # 활성 시트만

    header_row, project_cols_map, hidden_project_cols, dropped_duplicates = find_project_header_row_visible(ws)
    project_col_indices = sorted(project_cols_map.keys())
    if not project_col_indices:
        raise RuntimeError("보이는(Project 헤더가 있는) 프로젝트 열을 찾지 못했습니다.")

    label_col = column_index_from_string(label_col_letter)
    type_col = label_col + 1

    gi_block = find_section_block_by_top_sections(ws, label_col, "General inputs", TOP_LEVEL_SECTIONS)
    prod_block = find_section_block_by_top_sections(ws, label_col, "Production", TOP_LEVEL_SECTIONS)

    missing_sections = []
    if gi_block is None:
        missing_sections.append("General inputs")
    if prod_block is None:
        missing_sections.append("Production")

    projects: List[Dict[str, Any]] = []
    for col_idx in project_col_indices:
        pname = project_cols_map[col_idx]
        col_letter = get_column_letter(col_idx)

        general_inputs: Dict[str, Any] = {}
        production: Dict[str, Any] = {}

        if gi_block is not None:
            gi_start, gi_end = gi_block
            # General inputs는 null도 유지하는 편이 보통 유리
            general_inputs = extract_section_kv_for_project_flat(
                ws, gi_start, gi_end, label_col, type_col, col_idx,
                include_meta=include_meta,
                include_empty=True
            )

        if prod_block is not None:
            pr_start, pr_end = prod_block
            # Production은 값이 있는 것만 추출(빈칸은 제외) + 그룹화
            # production = extract_section_kv_for_project_grouped(
            #     ws, pr_start, pr_end, label_col, type_col, col_idx,
            #     include_meta=include_meta,
            #     include_empty=False
            # )
            production = extract_section_kv_for_project_grouped(
                ws, pr_start, pr_end, label_col, type_col, col_idx,
                include_meta=include_meta,
                include_empty=True   # ✅ 빈값도 포함 + "" 처리
            )

        projects.append(
            {
                "name": pname,
                "col": col_letter,
                "general_inputs": general_inputs,
                "production": production,
            }
        )

    return {
        "source": {
            "file": xlsx_path,
            "sheet": ws.title,
            "project_header_row": header_row,
            "label_col": label_col_letter,
            "projects_detected": [project_cols_map[c] for c in project_col_indices],
            "project_columns": [{"project": project_cols_map[c], "col": get_column_letter(c)} for c in project_col_indices],
            "hidden_project_columns": hidden_project_cols,
            "dropped_duplicate_projects": dropped_duplicates,
            "missing_sections": missing_sections,
        },
        "projects": projects,
    }


def main():
    ap = argparse.ArgumentParser(
        description="Extract 'General inputs' + 'Production' for ALL visible project columns from the active sheet into JSON."
    )
    ap.add_argument("xlsx", help="Input xlsx path")
    ap.add_argument("-c", "--label-col", default="B", help="Label column letter (default: B)")
    ap.add_argument("-o", "--out", default="projects_gi_prod_visible.json", help="Output JSON filename")
    ap.add_argument("--include-meta", action="store_true", help="Include type/row/col metadata")
    ap.add_argument("--indent", type=int, default=2, help="JSON indent (default: 2)")
    args = ap.parse_args()

    result = extract_all_projects_gi_and_production_visible(
        xlsx_path=args.xlsx,
        label_col_letter=args.label_col,
        include_meta=args.include_meta,
    )

    out_arg = Path(args.out)
    out_path = out_arg if out_arg.is_absolute() else (Path.cwd() / out_arg)
    if len(out_path.parts) == 1:
        out_path = Path.cwd() / "data" / out_path.name
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=args.indent)

    print(f"[OK] saved: {out_path}")
    print(f"     sheet: {result['source']['sheet']}")
    print(f"     projects: {len(result['projects'])}")
    if result["source"].get("hidden_project_columns"):
        print(f"     hidden_project_columns: {len(result['source']['hidden_project_columns'])}")
    if result["source"].get("dropped_duplicate_projects"):
        print(f"     dropped_duplicate_projects: {len(result['source']['dropped_duplicate_projects'])}")
    if result["source"].get("missing_sections"):
        print(f"     missing_sections: {result['source']['missing_sections']}")


if __name__ == "__main__":
    main()


"""
python ./pv_solar/test3/extract_general_inputs.py  "/ryan/11.airun/solar/workspaces_pv_solar/pv_solar_sample_docs/test2.xlsx"  -o general.json
"""
